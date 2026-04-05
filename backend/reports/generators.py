import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf(title, headers, rows, summary=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2A2626'),
        spaceAfter=12,
    )
    elements.append(Paragraph('PredictAI', title_style))
    elements.append(Paragraph(title, styles['Heading2']))
    elements.append(Paragraph(
        f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        styles['Normal'],
    ))
    elements.append(Spacer(1, 0.5 * cm))

    if summary:
        for key, val in summary.items():
            elements.append(Paragraph(f'{key}: {val}', styles['Normal']))
        elements.append(Spacer(1, 0.5 * cm))

    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2A2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4EFE6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2DAC9')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel(title, headers, rows, sheet_name='Relatorio'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2A2626', end_color='2A2626', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.append([title])
    ws.append([f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.append([])
    ws.append(headers)

    header_row = ws[4]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    alt_fill = PatternFill(start_color='F4EFE6', end_color='F4EFE6', fill_type='solid')
    for i, row in enumerate(rows):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
